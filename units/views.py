from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from .models import Unit, Booking, Report, Contract, Expense, UnitPricing, SpecialPricing, ProfitPercentage
from django.db.models import Sum, Count, Q
from datetime import datetime, timedelta
from calendar import monthrange
from django.views.decorators.cache import never_cache
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect
from django.views.decorators.http import require_POST
from django.db.models import Sum
import json
from io import BytesIO

def format_date_arabic(date_obj):
    """تحويل التاريخ إلى صيغة عربية مع التقويم الميلادي ويوم الأسبوع"""
    if not date_obj:
        return ""
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
        except:
            try:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d %H:%M:%S').date()
            except:
                return date_obj
    
    # أيام الأسبوع بالعربي
    weekdays_arabic = {
        0: 'الإثنين', 1: 'الثلاثاء', 2: 'الأربعاء', 3: 'الخميس',
        4: 'الجمعة', 5: 'السبت', 6: 'الأحد'
    }
    
    months_arabic = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    
    weekday_num = date_obj.weekday()  # 0=Monday, 6=Sunday
    weekday_name = weekdays_arabic[weekday_num]
    day = date_obj.day
    month = months_arabic[date_obj.month]
    year = date_obj.year
    
    return f"{weekday_name} {day} {month} {year}"

def setup_arabic_font():
    """تسجيل خط يدعم العربية - الأفضلية للخطوط التي تدعم العربية بشكل كامل"""
    try:
        # ترتيب الخطوط حسب الأفضلية (الأفضل أولاً)
        font_paths = [
            # Windows - Tahoma و Arial Unicode أفضل للعربية
            'C:/Windows/Fonts/tahoma.ttf',
            'C:/Windows/Fonts/tahomabd.ttf',  # Tahoma Bold
            'C:/Windows/Fonts/arialuni.ttf',  # Arial Unicode - يدعم العربية بشكل ممتاز
            'C:/Windows/Fonts/arial.ttf',
            'C:/Windows/Fonts/arialbd.ttf',    # Arial Bold
            'C:/Windows/Fonts/Times New Roman.ttf',
            # Linux - DejaVu Sans
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
            '/usr/share/fonts/TTF/DejaVuSans.ttf',
            '/usr/share/fonts/TTF/DejaVuSans-Bold.ttf',
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    # تسجيل الخط العادي
                    if 'bold' not in font_path.lower() and 'bd' not in font_path.lower():
                        pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                        # تسجيل الخط العريض أيضاً إن أمكن
                        try:
                            bold_path = font_path.replace('.ttf', 'bd.ttf')
                            if not os.path.exists(bold_path):
                                bold_path = font_path.replace('.ttf', 'Bold.ttf')
                            if os.path.exists(bold_path):
                                pdfmetrics.registerFont(TTFont('ArabicFont-Bold', bold_path))
                        except:
                            pass
                        return 'ArabicFont'
                    else:
                        # إذا كان خط عريض، نحتاج أيضاً للعادي
                        normal_path = font_path.replace('bd.ttf', '.ttf').replace('Bold.ttf', '.ttf')
                        if os.path.exists(normal_path):
                            pdfmetrics.registerFont(TTFont('ArabicFont', normal_path))
                            pdfmetrics.registerFont(TTFont('ArabicFont-Bold', font_path))
                            return 'ArabicFont'
                except Exception as e:
                    continue
        
        # إذا لم يتم العثور على خط يدعم العربية، استخدم Helvetica
        return 'Helvetica'
    except Exception:
        return 'Helvetica'

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import os

# محاولة استيراد مكتبات إعادة تشكيل النص العربي
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    ARABIC_SUPPORT = True
except ImportError:
    ARABIC_SUPPORT = False
    arabic_reshaper = None
    get_display = None

def reshape_arabic_text(text):
    """إعادة تشكيل النص العربي لعرضه بشكل صحيح من اليمين لليسار"""
    if not text or not ARABIC_SUPPORT:
        return text
    
    try:
        # إعادة تشكيل النص العربي
        reshaped_text = arabic_reshaper.reshape(text)
        # تحويل الاتجاه من اليمين إلى اليسار
        bidi_text = get_display(reshaped_text)
        return bidi_text
    except Exception:
        return text
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def home(request):
    """عرض الصفحة الرئيسية"""
    return render(request, 'index.html')

def services(request):
    """عرض صفحة الخدمات"""
    return render(request, 'services.html')

def policy(request):
    """عرض صفحة سياسة التشغيل"""
    return render(request, 'policy.html')

@login_required
@never_cache
def units(request):
    """عرض صفحة الوحدات (للمسجّلين فقط)"""
    units_qs = Unit.objects.filter(owner=request.user).prefetch_related('gallery_images')
    reports = Report.objects.filter(owner=request.user)[:20]
    contracts = Contract.objects.filter(owner=request.user)[:20]
    expenses = Expense.objects.filter(owner=request.user).select_related('unit')[:20]
    response = render(request, 'units.html', {
        'units': units_qs,
        'reports': reports,
        'contracts': contracts,
        'expenses': expenses,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response

@never_cache
def unit_bookings(request, unit_id):
    """إرجاع الحجوزات لوحدة معينة بصيغة JSON للتقويم"""
    unit = get_object_or_404(Unit, id=unit_id)
    bookings = Booking.objects.filter(unit=unit)
    expenses_qs = Expense.objects.filter(unit=unit)
    
    events = []
    total_booking_amount = 0.0
    for booking in bookings:
        # إنشاء حدث لكل يوم في فترة الحجز
        current_date = booking.start_date
        # عدد الأيام في الحجز الحالي (يشمل اليوم الأخير)
        total_days = (booking.end_date - booking.start_date).days + 1
        # حساب قيمة الحجز (سعر اليوم * عدد الأيام) أو المبالغ المسجلة
        cash_transfer_total = float((booking.cash_amount or 0) + (booking.transfer_amount or 0))
        if cash_transfer_total > 0:
            total_booking_amount += cash_transfer_total
        elif booking.price_per_day is not None:
            total_booking_amount += float(booking.price_per_day) * max(total_days, 1)
        while current_date <= booking.end_date:
            events.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'title': 'محجوز من المالك' if booking.is_owner_booking else 'محجوز',
                'color': '#dc3545',
                'price': float(booking.price_per_day) if booking.price_per_day is not None else None,
                'notes': booking.notes or '',
                'is_owner_booking': booking.is_owner_booking,
                'is_user_booking': (request.user.is_authenticated and booking.user_id == request.user.id)
            })
            current_date += timedelta(days=1)
    
    total_expenses = expenses_qs.aggregate(total=Sum('price'))['total'] or 0
    net_total = total_booking_amount - float(total_expenses)

    resp = JsonResponse({
        'events': events,
        'unit_name': unit.name,
        'total_booking_amount': round(total_booking_amount, 2),
        'total_expenses': round(float(total_expenses), 2),
        'net_total': round(net_total, 2),
    })
    resp['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp['Pragma'] = 'no-cache'
    return resp


@login_required
@never_cache
def dashboard(request):
    """لوحة بسيطة بعد تسجيل الدخول: نافبار + وحدات المالك + فوتر"""
    return redirect('units:units')


@login_required
@require_POST
@never_cache
def create_booking(request, unit_id):
    """إنشاء حجز ليوم واحد للمستخدم المسجل"""
    unit = get_object_or_404(Unit, id=unit_id)

    # قراءة البيانات JSON أو POST
    try:
        if request.content_type == 'application/json':
            payload = json.loads(request.body.decode('utf-8'))
        else:
            payload = request.POST
    except Exception:
        payload = {}

    date_str = payload.get('date')
    price = payload.get('price')
    notes = payload.get('notes')

    if not date_str:
        return JsonResponse({'error': 'التاريخ مطلوب'}, status=400)

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'تنسيق التاريخ غير صحيح'}, status=400)

    # إنشاء الحجز ليوم واحد
    booking = Booking(
        unit=unit,
        start_date=target_date,
        end_date=target_date,
        customer_name=request.user.get_full_name() or request.user.username,
        user=request.user,
        price_per_day=price if price not in (None, '',) else None,
        notes=notes or '',
        is_owner_booking=True
    )

    try:
        booking.save()
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'ok': True})


@login_required
@require_POST
@never_cache
def cancel_booking(request, unit_id):
    """إلغاء حجز يوم واحد (صلاحية للمالك أو الموظف)"""
    unit = get_object_or_404(Unit, id=unit_id)
    # السماح فقط لمالك الوحدة أو الموظف
    if not (request.user == unit.owner or request.user.is_staff):
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8')) if request.content_type == 'application/json' else request.POST
    except Exception:
        payload = request.POST

    date_str = payload.get('date')
    if not date_str:
        return JsonResponse({'error': 'التاريخ مطلوب'}, status=400)

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'تنسيق التاريخ غير صحيح'}, status=400)

    # إيجاد الحجز لليوم المحدد
    booking = Booking.objects.filter(unit=unit, start_date=target_date, end_date=target_date).first()
    if not booking:
        return JsonResponse({'error': 'لا يوجد حجز في هذا التاريخ'}, status=404)

    # الصلاحيات: موظف أو مالك الوحدة أو نفس صاحب الحجز (بحقل user أو بالاسم)
    is_owner = (request.user == unit.owner)
    is_staff = request.user.is_staff
    is_booking_owner_by_name = False
    is_booking_owner_by_user = (booking.user_id == request.user.id) if request.user.is_authenticated else False
    user_full_name = (request.user.get_full_name() or '').strip()
    user_username = (request.user.username or '').strip()
    if booking.customer_name:
        cn = booking.customer_name.strip()
        is_booking_owner_by_name = (cn == user_full_name) or (cn == user_username)

    if not (is_staff or is_owner or is_booking_owner_by_user or is_booking_owner_by_name):
        return JsonResponse({'error': 'غير مصرح لك بإلغاء هذا الحجز'}, status=403)

    try:
        booking.delete()
    except Exception as e:
        return JsonResponse({'error': f'تعذر إلغاء الحجز: {str(e)}'}, status=400)
    return JsonResponse({'ok': True})


@never_cache
def login_view(request):
    """تسجيل دخول منسق مع توجيه حسب الصلاحيات"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember = request.POST.get('remember') == 'on'
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            # ضبط مدة الجلسة: 48 ساعة إذا تم اختيار "تذكرني"، وإلا تنتهي عند إغلاق المتصفح
            if remember:
                request.session.set_expiry(172800)  # 48 ساعة
            else:
                request.session.set_expiry(0)  # عند إغلاق المتصفح
            # توجيه حسب الدور
            if user.is_staff:
                return HttpResponseRedirect('/admin/')
            return redirect('units:units')
        return render(request, 'registration/login.html', {
            'error': 'بيانات الدخول غير صحيحة'
        })
    return render(request, 'registration/login.html')


@login_required
def logout_view(request):
    logout(request)
    return redirect('units:login')


@staff_member_required
@never_cache
def payment_reports(request):
    """تقارير المدفوعات - عرض مفصل للكاش والتحويل مع فلترة"""
    from django.db.models import Q
    
    # الحصول على جميع الوحدات للقائمة
    all_units = Unit.objects.all().order_by('name')
    
    # معاملات الفلترة
    unit_id = request.GET.get('unit_id', '')
    report_type = request.GET.get('report_type', 'all')  # all, daily, weekly, monthly
    selected_date = request.GET.get('date', '')
    
    # البدء بجميع الحجوزات
    bookings = Booking.objects.all().select_related('unit', 'unit__owner')
    
    # فلترة حسب الوحدة
    if unit_id and unit_id != 'all':
        try:
            unit_filter = Unit.objects.get(id=unit_id)
            bookings = bookings.filter(unit=unit_filter)
            selected_unit = unit_filter
        except Unit.DoesNotExist:
            selected_unit = None
    else:
        selected_unit = None
    
    # فلترة حسب نوع التقرير
    if report_type == 'daily' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            bookings = bookings.filter(start_date=target_date)
        except ValueError:
            pass
    elif report_type == 'weekly' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            week_start = target_date - timedelta(days=target_date.weekday())
            week_end = week_start + timedelta(days=6)
            bookings = bookings.filter(start_date__gte=week_start, start_date__lte=week_end)
        except ValueError:
            pass
    elif report_type == 'monthly' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            month_start = target_date.replace(day=1)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
            bookings = bookings.filter(start_date__gte=month_start, start_date__lte=month_end)
        except ValueError:
            pass
    
    bookings = bookings.order_by('-start_date')
    
    # حساب الإجماليات
    total_cash = bookings.aggregate(Sum('cash_amount'))['cash_amount__sum'] or 0
    total_transfer = bookings.aggregate(Sum('transfer_amount'))['transfer_amount__sum'] or 0
    total_all = total_cash + total_transfer
    
    # فصل الحجوزات حسب نوع الدفع
    cash_bookings = bookings.filter(cash_amount__gt=0).order_by('-start_date')
    transfer_bookings = bookings.filter(transfer_amount__gt=0).order_by('-start_date')
    
    context = {
        'bookings': bookings,
        'cash_bookings': cash_bookings,
        'transfer_bookings': transfer_bookings,
        'total_cash': total_cash,
        'total_transfer': total_transfer,
        'total_all': total_all,
        'all_units': all_units,
        'selected_unit_id': unit_id or 'all',
        'selected_unit': selected_unit,
        'selected_report_type': report_type,
        'selected_date': selected_date,
    }
    
    return render(request, 'admin/payment_reports.html', context)


@staff_member_required
@never_cache
def payment_reports_pdf(request):
    """تصدير تقارير المدفوعات كـ PDF مع فلترة"""
    # نفس منطق الفلترة من payment_reports
    unit_id = request.GET.get('unit_id', '')
    report_type = request.GET.get('report_type', 'all')
    selected_date = request.GET.get('date', '')
    
    bookings = Booking.objects.all().select_related('unit', 'unit__owner')
    
    # فلترة حسب الوحدة
    if unit_id and unit_id != 'all':
        try:
            unit_filter = Unit.objects.get(id=unit_id)
            bookings = bookings.filter(unit=unit_filter)
        except Unit.DoesNotExist:
            pass
    
    # فلترة حسب نوع التقرير
    if report_type == 'daily' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            bookings = bookings.filter(start_date=target_date)
        except ValueError:
            pass
    elif report_type == 'weekly' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            week_start = target_date - timedelta(days=target_date.weekday())
            week_end = week_start + timedelta(days=6)
            bookings = bookings.filter(start_date__gte=week_start, start_date__lte=week_end)
        except ValueError:
            pass
    elif report_type == 'monthly' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            month_start = target_date.replace(day=1)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
            bookings = bookings.filter(start_date__gte=month_start, start_date__lte=month_end)
        except ValueError:
            pass
    
    bookings = bookings.order_by('-start_date')
    
    total_cash = bookings.aggregate(Sum('cash_amount'))['cash_amount__sum'] or 0
    total_transfer = bookings.aggregate(Sum('transfer_amount'))['transfer_amount__sum'] or 0
    total_all = total_cash + total_transfer
    
    # تسجيل خط يدعم العربية
    arabic_font = setup_arabic_font()
    if arabic_font == 'ArabicFont':
        # محاولة استخدام الخط العريض المسجل
        try:
            if 'ArabicFont-Bold' in pdfmetrics.getRegisteredFontNames():
                arabic_font_bold = 'ArabicFont-Bold'
            else:
                arabic_font_bold = 'ArabicFont'
        except:
            arabic_font_bold = 'ArabicFont'
    else:
        arabic_font_bold = 'Helvetica-Bold'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=1*cm,  # هوامش أقل = مساحة أكثر
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm,
        title="تقارير المدفوعات"
    )
    elements = []
    
    styles = getSampleStyleSheet()
    
    # نمط العنوان الرئيسي
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#000000'),
        spaceAfter=15,
        alignment=1,  # center للعناوين
        fontName=arabic_font_bold,
        encoding='utf-8',
        rightIndent=0,
        leftIndent=0
    )
    
    # نمط العنوان الفرعي
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#000000'),
        spaceAfter=10,
        alignment=1,  # center للعناوين
        fontName=arabic_font,
        encoding='utf-8',
        rightIndent=0,
        leftIndent=0
    )
    
    # العنوان الرئيسي
    main_title = Paragraph(reshape_arabic_text("تقارير المدفوعات"), title_style)
    elements.append(main_title)
    
    # معلومات الشركة
    company_info = Paragraph(reshape_arabic_text("القمة العقارية"), subtitle_style)
    elements.append(company_info)
    elements.append(Spacer(1, 0.3*cm))
    
    # معلومات الفلترة
    filter_info = []
    if unit_id and unit_id != 'all':
        try:
            unit_name = Unit.objects.get(id=unit_id).name
            filter_info.append(reshape_arabic_text(f"الوحدة: {unit_name}"))
        except:
            pass
    if report_type == 'daily' and selected_date:
        filter_info.append(reshape_arabic_text(f"التقرير: يومي - {format_date_arabic(selected_date)}"))
    elif report_type == 'weekly' and selected_date:
        filter_info.append(reshape_arabic_text(f"التقرير: أسبوعي - {format_date_arabic(selected_date)}"))
    elif report_type == 'monthly' and selected_date:
        filter_info.append(reshape_arabic_text(f"التقرير: شهري - {format_date_arabic(selected_date)}"))
    elif report_type == 'all':
        filter_info.append(reshape_arabic_text("التقرير: جميع الحجوزات"))
    
    # تاريخ التصدير
    export_date = format_date_arabic(datetime.now().date())
    filter_info.append(reshape_arabic_text(f"تاريخ التصدير: {export_date}"))
    
    if filter_info:
        filter_text = " | ".join(filter_info)
        filter_para = Paragraph(filter_text, subtitle_style)
        elements.append(filter_para)
    
    elements.append(Spacer(1, 0.5*cm))
    
    # استخدام Paragraph للنصوص العربية - تنسيق أفضل للطباعة
    table_style = ParagraphStyle(
        'TableStyle',
        parent=styles['Normal'],
        fontSize=10,  # حجم أكبر للوضوح
        alignment=1,  # CENTER
        fontName=arabic_font,
        encoding='utf-8',
        rightIndent=0,
        leftIndent=0,
        spaceBefore=0,
        spaceAfter=0,
        leading=12  # مسافة أكبر بين الأسطر
    )
    
    # نمط الجدول للعنوان - مطابق للوحة التحكم
    table_header_style = ParagraphStyle(
        'TableHeaderStyle',
        parent=styles['Normal'],
        fontSize=10,  # حجم مناسب
        alignment=1,  # CENTER - العناوين في الوسط
        fontName=arabic_font_bold,
        encoding='utf-8',
        rightIndent=0,
        leftIndent=0,
        leading=12  # المسافة بين الأسطر
    )
    
    headers = [Paragraph(reshape_arabic_text('الوحدة'), table_header_style), 
               Paragraph(reshape_arabic_text('التاريخ'), table_header_style), 
               Paragraph(reshape_arabic_text('اسم العميل'), table_header_style), 
               Paragraph(reshape_arabic_text('رقم الهاتف'), table_header_style), 
               Paragraph(reshape_arabic_text('الكاش'), table_header_style), 
               Paragraph(reshape_arabic_text('التحويل'), table_header_style), 
               Paragraph(reshape_arabic_text('الإجمالي'), table_header_style)]
    
    data = [headers]
    
    for booking in bookings:
        unit_name = Paragraph(reshape_arabic_text(booking.unit.name), table_style)
        date_str = Paragraph(reshape_arabic_text(format_date_arabic(booking.start_date)), table_style)
        customer_name = Paragraph(reshape_arabic_text(booking.customer_name or '-'), table_style)
        customer_phone = Paragraph(reshape_arabic_text(booking.customer_phone or '-'), table_style)
        cash = f"{float(booking.cash_amount):,.2f}" if booking.cash_amount else "0.00"
        transfer = f"{float(booking.transfer_amount):,.2f}" if booking.transfer_amount else "0.00"
        total = f"{float(booking.total_amount):,.2f}"
        cash_amount = Paragraph(reshape_arabic_text(cash + ' ر.س'), table_style)
        transfer_amount = Paragraph(reshape_arabic_text(transfer + ' ر.س'), table_style)
        total_amount = Paragraph(reshape_arabic_text(total + ' ر.س'), table_style)
        data.append([unit_name, date_str, customer_name, customer_phone, 
                    cash_amount, transfer_amount, total_amount])
    
    total_row = [Paragraph(reshape_arabic_text('الإجمالي'), table_style), 
                 Paragraph('', table_style), 
                 Paragraph('', table_style), 
                 Paragraph('', table_style), 
                 Paragraph(reshape_arabic_text(f"{float(total_cash):,.2f} ر.س"), table_style), 
                 Paragraph(reshape_arabic_text(f"{float(total_transfer):,.2f} ر.س"), table_style), 
                 Paragraph(reshape_arabic_text(f"{float(total_all):,.2f} ر.س"), table_style)]
    data.append(total_row)
    
    # عرض الأعمدة حسب المحتوى - بدون ضغط
    available_width = A4[0] - (1*cm * 2)  # هوامش أقل = مساحة أكثر
    
    # توزيع الأعمدة بشكل متساوٍ على العرض المتاح - بدون ضغط
    num_cols = 7
    col_width = available_width / num_cols  # كل عمود يأخذ نفس العرض
    
    # لكن نعطي مساحة أكبر للأعمدة التي تحتاجها
    col_widths = [
        col_width * 1.0,  # الوحدة
        col_width * 1.5,  # التاريخ - مساحة أكبر بكثير ليوم الأسبوع الكامل
        col_width * 1.2,  # اسم العميل - مساحة جيدة
        col_width * 1.4,  # رقم الهاتف - مساحة أكبر بكثير ليكون كامل
        col_width * 0.9,  # الكاش
        col_width * 0.9,  # التحويل
        col_width * 0.9,  # الإجمالي
    ]
    
    # ضبط العرض الإجمالي
    total_width = sum(col_widths)
    if total_width > available_width:
        ratio = available_width / total_width
        col_widths = [w * ratio for w in col_widths]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    table.setStyle(TableStyle([
        # رأس الجدول - ألوان الهوية (تدرج البيج)
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#a89078')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # كل النصوص في الوسط (مثل لوحة التحكم)
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), arabic_font_bold if arabic_font != 'Helvetica' else 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),  # حجم أكبر للوضوح
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('LEFTPADDING', (0, 0), (-1, 0), 12),  # مساحة أكثر
        ('RIGHTPADDING', (0, 0), (-1, 0), 12),
        # بيانات الجدول - النصوص سوداء
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#000000')),
        ('FONTNAME', (0, 1), (-1, -2), arabic_font),
        ('FONTSIZE', (0, 1), (-1, -2), 10),  # حجم أكبر للوضوح
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#fefcf8')]),
        ('TOPPADDING', (0, 1), (-1, -2), 12),  # مساحة أكثر للمحتوى
        ('BOTTOMPADDING', (0, 1), (-1, -2), 12),
        ('LEFTPADDING', (0, 1), (-1, -2), 12),  # مساحة أكثر
        ('RIGHTPADDING', (0, 1), (-1, -2), 12),
        # صف الإجمالي - ألوان الهوية (بيج ذهبي)
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#a89078')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), arabic_font_bold if arabic_font != 'Helvetica' else 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
        ('LEFTPADDING', (0, -1), (-1, -1), 12),
        ('RIGHTPADDING', (0, -1), (-1, -1), 12),
        # الحدود - مطابق للوحة التحكم (1px solid #ddd)
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),  # 1px solid #e2e8f0
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))
    
    # ملخص الإجماليات
    summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=5,
        alignment=0,  # right for RTL
        fontName=arabic_font_bold,
        encoding='utf-8'
    )
    
    summary_text = f"""
    <b>ملخص الإجماليات:</b><br/>
    • إجمالي المدفوعات نقداً: {float(total_cash):,.2f} ر.س<br/>
    • إجمالي المدفوعات تحويل: {float(total_transfer):,.2f} ر.س<br/>
    • <b>الإجمالي الكلي: {float(total_all):,.2f} ر.س</b><br/>
    • عدد الحجوزات: {bookings.count()} حجز
    """
    
    summary = Paragraph(reshape_arabic_text(summary_text), summary_style)
    elements.append(summary)
    
    # بناء PDF
    doc.build(elements)
    buffer.seek(0)
    
    # قراءة محتوى الـ buffer
    pdf_content = buffer.read()
    buffer.close()
    
    # إنشاء الرد مع التأكد من أن الملف PDF
    response = HttpResponse(
        pdf_content,
        content_type='application/pdf'
    )
    response['Content-Disposition'] = f'attachment; filename="payment_reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response['Content-Length'] = str(len(pdf_content))
    
    # التأكد من أن المتصفح يعرف أن هذا ملف PDF وليس Word
    response['X-Content-Type-Options'] = 'nosniff'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response


@staff_member_required
@never_cache
def payment_reports_excel(request):
    """تصدير تقارير المدفوعات كـ Excel مع فلترة"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel export requires openpyxl. Install it: pip install openpyxl", status=500)
    
    # نفس منطق الفلترة
    unit_id = request.GET.get('unit_id', '')
    report_type = request.GET.get('report_type', 'all')
    selected_date = request.GET.get('date', '')
    
    bookings = Booking.objects.all().select_related('unit', 'unit__owner')
    
    # فلترة حسب الوحدة
    if unit_id and unit_id != 'all':
        try:
            unit_filter = Unit.objects.get(id=unit_id)
            bookings = bookings.filter(unit=unit_filter)
        except Unit.DoesNotExist:
            pass
    
    # فلترة حسب نوع التقرير
    if report_type == 'daily' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            bookings = bookings.filter(start_date=target_date)
        except ValueError:
            pass
    elif report_type == 'weekly' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            week_start = target_date - timedelta(days=target_date.weekday())
            week_end = week_start + timedelta(days=6)
            bookings = bookings.filter(start_date__gte=week_start, start_date__lte=week_end)
        except ValueError:
            pass
    elif report_type == 'monthly' and selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            month_start = target_date.replace(day=1)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
            bookings = bookings.filter(start_date__gte=month_start, start_date__lte=month_end)
        except ValueError:
            pass
    
    bookings = bookings.order_by('-start_date')
    
    total_cash = bookings.aggregate(Sum('cash_amount'))['cash_amount__sum'] or 0
    total_transfer = bookings.aggregate(Sum('transfer_amount'))['transfer_amount__sum'] or 0
    total_all = total_cash + total_transfer
    
    wb = Workbook()
    ws = wb.active
    ws.title = "تقارير المدفوعات"
    
    # الأنماط - ألوان الهوية (تدرج البيج)
    header_fill = PatternFill(start_color="a89078", end_color="8b7765", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="a89078", end_color="a89078", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # بناء العنوان مع معلومات الفلترة
    title_text = 'تقارير المدفوعات - القمة العقارية'
    if unit_id and unit_id != 'all':
        try:
            unit_name = Unit.objects.get(id=unit_id).name
            title_text += f' | الوحدة: {unit_name}'
        except:
            pass
    if report_type == 'daily' and selected_date:
        title_text += f' | يومي - {format_date_arabic(selected_date)}'
    elif report_type == 'weekly' and selected_date:
        title_text += f' | أسبوعي - {format_date_arabic(selected_date)}'
    elif report_type == 'monthly' and selected_date:
        title_text += f' | شهري - {format_date_arabic(selected_date)}'
    
    # العنوان
    ws.merge_cells('A1:G1')
    ws['A1'] = title_text
    title_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0)
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = title_alignment
    
    # رأس الجدول
    headers = ['الوحدة', 'التاريخ', 'اسم العميل', 'رقم الهاتف', 'الكاش', 'التحويل', 'الإجمالي']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # البيانات
    row = 4
    for booking in bookings:
        # النصوص العربية
        cell1 = ws.cell(row=row, column=1, value=booking.unit.name)
        cell1.border = border
        cell1.alignment = center_alignment
        
        cell2 = ws.cell(row=row, column=2, value=format_date_arabic(booking.start_date))
        cell2.border = border
        cell2.alignment = center_alignment
        
        cell3 = ws.cell(row=row, column=3, value=booking.customer_name or '-')
        cell3.border = border
        cell3.alignment = center_alignment
        
        cell4 = ws.cell(row=row, column=4, value=booking.customer_phone or '-')
        cell4.border = border
        cell4.alignment = center_alignment
        
        # الأرقام مع تنسيق
        cash_value = float(booking.cash_amount or 0)
        cell5 = ws.cell(row=row, column=5, value=f"{cash_value:,.2f} ر.س")
        cell5.border = border
        cell5.alignment = center_alignment
        
        transfer_value = float(booking.transfer_amount or 0)
        cell6 = ws.cell(row=row, column=6, value=f"{transfer_value:,.2f} ر.س")
        cell6.border = border
        cell6.alignment = center_alignment
        
        total_value = float(booking.total_amount)
        cell7 = ws.cell(row=row, column=7, value=f"{total_value:,.2f} ر.س")
        cell7.border = border
        cell7.alignment = center_alignment
        
        row += 1
    
    # الصف الأخير - الإجماليات
    cell1 = ws.cell(row=row, column=1, value='الإجمالي')
    cell1.fill = total_fill
    cell1.font = total_font
    cell1.alignment = center_alignment
    cell1.border = border
    
    for col in [2, 3, 4]:
        cell = ws.cell(row=row, column=col, value='')
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = center_alignment
        cell.border = border
    
    cell5 = ws.cell(row=row, column=5, value=f"{float(total_cash):,.2f} ر.س")
    cell5.fill = total_fill
    cell5.font = total_font
    cell5.alignment = center_alignment
    cell5.border = border
    
    cell6 = ws.cell(row=row, column=6, value=f"{float(total_transfer):,.2f} ر.س")
    cell6.fill = total_fill
    cell6.font = total_font
    cell6.alignment = center_alignment
    cell6.border = border
    
    cell7 = ws.cell(row=row, column=7, value=f"{float(total_all):,.2f} ر.س")
    cell7.fill = total_fill
    cell7.font = total_font
    cell7.alignment = center_alignment
    cell7.border = border
    
    # تعديل عرض الأعمدة
    ws.column_dimensions['A'].width = 20  # الوحدة
    ws.column_dimensions['B'].width = 25  # التاريخ (مع يوم الأسبوع)
    ws.column_dimensions['C'].width = 18  # اسم العميل
    ws.column_dimensions['D'].width = 15  # رقم الهاتف
    ws.column_dimensions['E'].width = 15  # الكاش
    ws.column_dimensions['F'].width = 15  # التحويل
    ws.column_dimensions['G'].width = 15  # الإجمالي
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payment_reports_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


@login_required
@never_cache
def unit_expenses(request, unit_id):
    """عرض مصروفات وحدة معينة"""
    unit = get_object_or_404(Unit, id=unit_id, owner=request.user)
    expenses = Expense.objects.filter(unit=unit, owner=request.user).order_by('-created_at')
    
    # حساب الإجمالي
    total_expenses = expenses.aggregate(Sum('price'))['price__sum'] or 0
    
    context = {
        'unit': unit,
        'expenses': expenses,
        'total_expenses': total_expenses,
    }
    
    response = render(request, 'unit_expenses.html', context)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


@login_required
@never_cache
def expense_detail(request, expense_id):
    """عرض تفاصيل مصروف معين مع الفاتورة"""
    expense = get_object_or_404(Expense, id=expense_id, owner=request.user)
    
    context = {
        'expense': expense,
    }
    
    response = render(request, 'expense_detail.html', context)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


@login_required
@never_cache
def unit_pricing(request, unit_id):
    """عرض أسعار تأجير وحدة معينة"""
    unit = get_object_or_404(Unit, id=unit_id, owner=request.user)
    
    # جلب جميع أسعار الوحدة الأساسية (وسط الأسبوع: خميس، جمعة، سبت)
    pricing_list = UnitPricing.objects.filter(unit=unit, day_of_week__in=[3, 4, 5]).order_by('day_of_week')
    
    # إنشاء قائمة بأسعار وسط الأسبوع
    weekdays_info = [
        (3, 'الخميس'),
        (4, 'الجمعة'),
        (5, 'السبت'),
    ]
    
    pricing_data = []
    pricing_dict = {p.day_of_week: p for p in pricing_list}
    
    for day_num, day_name in weekdays_info:
        pricing_obj = pricing_dict.get(day_num)
        pricing_data.append({
            'day_num': day_num,
            'day_name': day_name,
            'price': pricing_obj.price if pricing_obj else None,
            'exists': pricing_obj is not None
        })
    
    # جلب الأسعار الخاصة (عيد الفطر، عيد الأضحى، إجازات)
    special_pricing_list = SpecialPricing.objects.filter(unit=unit).order_by('pricing_type', 'night_number')
    
    # تنظيم الأسعار الخاصة حسب النوع
    eid_al_fitr_prices = {}
    eid_al_adha_prices = {}
    holiday_prices = {}
    
    for sp in special_pricing_list:
        price_data = {
            'night_number': sp.night_number,
            'night_name': sp.get_night_number_display_ar(),
            'price': sp.price
        }
        
        if sp.pricing_type == 'eid_al_fitr':
            eid_al_fitr_prices[sp.night_number] = price_data
        elif sp.pricing_type == 'eid_al_adha':
            eid_al_adha_prices[sp.night_number] = price_data
        elif sp.pricing_type == 'holiday':
            holiday_prices[sp.night_number] = price_data
    
    # إنشاء قوائم مرتبة للأسعار الخاصة (1-6)
    NIGHT_CHOICES_DICT = dict(SpecialPricing.NIGHT_CHOICES)
    def create_night_list(prices_dict):
        nights = []
        for night_num in range(1, 7):
            if night_num in prices_dict:
                nights.append(prices_dict[night_num])
            else:
                nights.append({
                    'night_number': night_num,
                    'night_name': NIGHT_CHOICES_DICT.get(night_num, f'الليلة {night_num}'),
                    'price': None
                })
        return nights
    
    eid_al_fitr_nights = create_night_list(eid_al_fitr_prices)
    eid_al_adha_nights = create_night_list(eid_al_adha_prices)
    holiday_nights = create_night_list(holiday_prices)
    
    context = {
        'unit': unit,
        'pricing_data': pricing_data,
        'eid_al_fitr_nights': eid_al_fitr_nights,
        'eid_al_adha_nights': eid_al_adha_nights,
        'holiday_nights': holiday_nights,
    }
    
    response = render(request, 'unit_pricing.html', context)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


@login_required
@never_cache
def unit_gallery(request, unit_id):
    """عرض صور وحدة معينة"""
    unit = get_object_or_404(Unit, id=unit_id, owner=request.user)
    images = unit.gallery_images.all()
    
    context = {
        'unit': unit,
        'images': images,
    }
    
    response = render(request, 'unit_gallery.html', context)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


@staff_member_required
@never_cache
def profits_view(request):
    """عرض الأرباح والرسوم البيانية للمديرين"""
    from collections import defaultdict
    
    # حساب الأرباح لكل وحدة
    units = Unit.objects.all().order_by('name')
    profits_data = []
    
    for unit in units:
        # حساب إجمالي الحجوزات
        bookings = Booking.objects.filter(unit=unit)
        total_booking_amount = 0
        
        for booking in bookings:
            cash_transfer_total = float((booking.cash_amount or 0) + (booking.transfer_amount or 0))
            if cash_transfer_total > 0:
                total_booking_amount += cash_transfer_total
            elif booking.price_per_day is not None:
                total_days = (booking.end_date - booking.start_date).days + 1
                total_booking_amount += float(booking.price_per_day) * max(total_days, 1)
        
        # حساب إجمالي المصروفات
        expenses = Expense.objects.filter(unit=unit)
        total_expenses = float(expenses.aggregate(Sum('price'))['price__sum'] or 0)
        
        # حساب الصافي
        net_total = total_booking_amount - total_expenses
        
        # الحصول على نسبة الأرباح من مالك الوحدة
        percentage = 50  # افتراضي 50%
        if unit.owner:
            try:
                profit_percentage_obj = unit.owner.profit_percentage
                percentage = profit_percentage_obj.percentage
            except ProfitPercentage.DoesNotExist:
                percentage = 50  # افتراضي 50%
        
        # حساب الأرباح
        profit = (net_total * percentage) / 100
        
        profits_data.append({
            'unit': unit,
            'owner': unit.owner,
            'total_bookings': total_booking_amount,
            'total_expenses': total_expenses,
            'net_total': net_total,
            'percentage': percentage,
            'profit': profit,
        })
    
    # بيانات الرسوم البيانية - أعلى الوحدات بالحجوزات
    today = datetime.now().date()
    current_month_start = today.replace(day=1)
    current_year_start = today.replace(month=1, day=1)
    
    # الحجوزات خلال الشهر
    monthly_bookings_raw = Booking.objects.filter(
        start_date__gte=current_month_start,
        start_date__lte=today
    ).select_related('unit')
    
    monthly_bookings_dict = {}
    for booking in monthly_bookings_raw:
        unit_name = booking.unit.name
        if unit_name not in monthly_bookings_dict:
            monthly_bookings_dict[unit_name] = {'booking_count': 0, 'total_amount': 0}
        monthly_bookings_dict[unit_name]['booking_count'] += 1
        cash_transfer_total = float((booking.cash_amount or 0) + (booking.transfer_amount or 0))
        if cash_transfer_total > 0:
            monthly_bookings_dict[unit_name]['total_amount'] += cash_transfer_total
        elif booking.price_per_day is not None:
            total_days = (booking.end_date - booking.start_date).days + 1
            monthly_bookings_dict[unit_name]['total_amount'] += float(booking.price_per_day) * max(total_days, 1)
    
    monthly_bookings = [{'unit__name': k, 'booking_count': v['booking_count'], 'total_amount': v['total_amount']} 
                        for k, v in sorted(monthly_bookings_dict.items(), key=lambda x: x[1]['booking_count'], reverse=True)[:10]]
    
    # الحجوزات خلال السنة
    yearly_bookings_raw = Booking.objects.filter(
        start_date__gte=current_year_start,
        start_date__lte=today
    ).select_related('unit')
    
    yearly_bookings_dict = {}
    for booking in yearly_bookings_raw:
        unit_name = booking.unit.name
        if unit_name not in yearly_bookings_dict:
            yearly_bookings_dict[unit_name] = {'booking_count': 0, 'total_amount': 0}
        yearly_bookings_dict[unit_name]['booking_count'] += 1
        cash_transfer_total = float((booking.cash_amount or 0) + (booking.transfer_amount or 0))
        if cash_transfer_total > 0:
            yearly_bookings_dict[unit_name]['total_amount'] += cash_transfer_total
        elif booking.price_per_day is not None:
            total_days = (booking.end_date - booking.start_date).days + 1
            yearly_bookings_dict[unit_name]['total_amount'] += float(booking.price_per_day) * max(total_days, 1)
    
    yearly_bookings = [{'unit__name': k, 'booking_count': v['booking_count'], 'total_amount': v['total_amount']} 
                       for k, v in sorted(yearly_bookings_dict.items(), key=lambda x: x[1]['booking_count'], reverse=True)[:10]]
    
    # المصروفات خلال الشهر
    monthly_expenses = Expense.objects.filter(
        created_at__date__gte=current_month_start,
        created_at__date__lte=today
    ).values('unit__name').annotate(
        total_expenses=Sum('price')
    ).order_by('-total_expenses')[:10]
    
    # المصروفات خلال السنة
    yearly_expenses = Expense.objects.filter(
        created_at__date__gte=current_year_start,
        created_at__date__lte=today
    ).values('unit__name').annotate(
        total_expenses=Sum('price')
    ).order_by('-total_expenses')[:10]
    
    context = {
        'profits_data': profits_data,
        'monthly_bookings': monthly_bookings,
        'yearly_bookings': yearly_bookings,
        'monthly_expenses': list(monthly_expenses),
        'yearly_expenses': list(yearly_expenses),
    }
    
    response = render(request, 'admin/profits.html', context)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


@staff_member_required
@never_cache
def profits_pdf(request):
    """تصدير تقرير الأرباح كـ PDF"""
    # حساب الأرباح لكل وحدة (نفس منطق profits_view)
    units = Unit.objects.all().order_by('name')
    profits_data = []
    
    for unit in units:
        # حساب إجمالي الحجوزات
        bookings = Booking.objects.filter(unit=unit)
        total_booking_amount = 0
        
        for booking in bookings:
            cash_transfer_total = float((booking.cash_amount or 0) + (booking.transfer_amount or 0))
            if cash_transfer_total > 0:
                total_booking_amount += cash_transfer_total
            elif booking.price_per_day is not None:
                total_days = (booking.end_date - booking.start_date).days + 1
                total_booking_amount += float(booking.price_per_day) * max(total_days, 1)
        
        # حساب إجمالي المصروفات
        expenses = Expense.objects.filter(unit=unit)
        total_expenses = float(expenses.aggregate(Sum('price'))['price__sum'] or 0)
        
        # حساب الصافي
        net_total = total_booking_amount - total_expenses
        
        # الحصول على نسبة الأرباح من مالك الوحدة
        percentage = 50  # افتراضي 50%
        if unit.owner:
            try:
                profit_percentage_obj = unit.owner.profit_percentage
                percentage = profit_percentage_obj.percentage
            except ProfitPercentage.DoesNotExist:
                percentage = 50  # افتراضي 50%
        
        # حساب الأرباح
        profit = (net_total * percentage) / 100
        
        profits_data.append({
            'unit': unit,
            'owner': unit.owner,
            'total_bookings': total_booking_amount,
            'total_expenses': total_expenses,
            'net_total': net_total,
            'percentage': percentage,
            'profit': profit,
        })
    
    # إنشاء PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    # تسجيل الخط العربي
    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'static/fonts/Cairo-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('ArabicBold', 'static/fonts/Cairo-Bold.ttf'))
        arabic_font = 'Arabic'
        arabic_font_bold = 'ArabicBold'
    except:
        arabic_font = 'Helvetica'
        arabic_font_bold = 'Helvetica-Bold'
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#8b7765'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=arabic_font_bold
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        fontName=arabic_font_bold,
        textColor=colors.white,
        backColor=colors.HexColor('#8b7765')
    )
    
    table_style = ParagraphStyle(
        'TableStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        fontName=arabic_font
    )
    
    # العنوان
    title = Paragraph(reshape_arabic_text('تقرير الأرباح'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # رؤوس الجدول
    headers = [
        Paragraph(reshape_arabic_text('الوحدة'), table_header_style),
        Paragraph(reshape_arabic_text('المالك'), table_header_style),
        Paragraph(reshape_arabic_text('إجمالي الحجوزات'), table_header_style),
        Paragraph(reshape_arabic_text('إجمالي المصروفات'), table_header_style),
        Paragraph(reshape_arabic_text('الصافي'), table_header_style),
        Paragraph(reshape_arabic_text('نسبة الأرباح'), table_header_style),
        Paragraph(reshape_arabic_text('الأرباح'), table_header_style)
    ]
    
    data = [headers]
    
    # إجماليات
    total_bookings = 0
    total_expenses = 0
    total_net = 0
    total_profit = 0
    
    for item in profits_data:
        unit_name = Paragraph(reshape_arabic_text(item['unit'].name), table_style)
        owner_name = Paragraph(reshape_arabic_text(item['owner'].username if item['owner'] else '-'), table_style)
        bookings = Paragraph(reshape_arabic_text(f"{item['total_bookings']:,.2f} ر.س"), table_style)
        expenses = Paragraph(reshape_arabic_text(f"{item['total_expenses']:,.2f} ر.س"), table_style)
        net = Paragraph(reshape_arabic_text(f"{item['net_total']:,.2f} ر.س"), table_style)
        percentage = Paragraph(reshape_arabic_text(f"{item['percentage']}%"), table_style)
        profit = Paragraph(reshape_arabic_text(f"{item['profit']:,.2f} ر.س"), table_style)
        
        data.append([unit_name, owner_name, bookings, expenses, net, percentage, profit])
        
        total_bookings += item['total_bookings']
        total_expenses += item['total_expenses']
        total_net += item['net_total']
        total_profit += item['profit']
    
    # صف الإجمالي
    total_row = [
        Paragraph(reshape_arabic_text('الإجمالي'), table_style),
        Paragraph('', table_style),
        Paragraph(reshape_arabic_text(f"{total_bookings:,.2f} ر.س"), table_style),
        Paragraph(reshape_arabic_text(f"{total_expenses:,.2f} ر.س"), table_style),
        Paragraph(reshape_arabic_text(f"{total_net:,.2f} ر.س"), table_style),
        Paragraph('', table_style),
        Paragraph(reshape_arabic_text(f"{total_profit:,.2f} ر.س"), table_style)
    ]
    data.append(total_row)
    
    # إنشاء الجدول
    available_width = A4[0] - (1*cm * 2)
    num_cols = 7
    col_width = available_width / num_cols
    
    table = Table(data, colWidths=[col_width] * num_cols)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b7765')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), arabic_font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('FONTNAME', (0, 1), (-1, -2), arabic_font),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, -1), (-1, -1), arabic_font_bold),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="profits_report.pdf"'
    return response


